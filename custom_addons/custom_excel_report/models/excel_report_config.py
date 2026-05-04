from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import date, datetime
from copy import copy

class ExcelReportConfig(models.Model):
    _name = 'excel.report.config'
    _description = 'Cấu hình báo cáo Excel'

    name = fields.Char(string='Tên báo cáo', required=True)
    store_procedure_name = fields.Char(string='Tên Function/Procedure', required=True)
    template_file = fields.Binary(string='File Excel Mẫu', required=True)
    file_name = fields.Char(string='Tên file')
    start_row = fields.Integer(string='Dòng chứa thẻ Mapping dữ liệu', default=8)
    
    param_ids = fields.One2many('excel.report.param', 'config_id', string='Tham số báo cáo')
    sql_column_ids = fields.One2many('excel.report.column', 'config_id', string='Danh sách cột SQL')
    
    group_by_column_ids = fields.Many2many(
        'excel.report.column', 
        'report_config_column_rel', 
        'config_id', 'column_id',
        string='Nhóm theo (Phân cấp)',
    )

    sum_column_ids = fields.Many2many(
        'excel.report.column',
        'report_config_sum_col_rel',
        'config_id', 'column_id',
        string='Cột cần tính Tổng (Sum)',
        domain="[('config_id', '=', id)]"
    )

    def action_get_columns(self):
        """Lấy danh sách cột từ Procedure"""
        self.ensure_one()
        if not self.store_procedure_name:
            raise UserError(_("Vui lòng nhập tên Function/Procedure."))
        
        params = []
        placeholders = []
        for p in self.param_ids.sorted('sequence'):
            if p.param_type == 'date':
                params.append(date.today()); placeholders.append("%s::date")
            elif p.param_type == 'char':
                params.append(""); placeholders.append("%s::varchar")
            else:
                params.append(0); placeholders.append("%s::integer")

        query = f"SELECT * FROM {self.store_procedure_name}({', '.join(placeholders)}) LIMIT 0"
        try:
            self.env.cr.execute(query, params)
            columns = [desc[0] for desc in self.env.cr.description]
            self.sql_column_ids.unlink()
            self.write({'sql_column_ids': [(0, 0, {'name': col}) for col in columns]})
        except Exception as e:
            raise UserError(_("Lỗi SQL: %s") % str(e))

    def action_export_report(self):
        self.ensure_one()
        
        # 1. Chuẩn bị tham số
        params = []
        param_dict = {}
        for p in self.param_ids.sorted('sequence'):
            val = p.value_date if p.param_type == 'date' else (p.value_char or "") if p.param_type == 'char' else p.value_integer
            params.append(val)
            display_val = val.strftime('%d/%m/%Y') if isinstance(val, (date, datetime)) else str(val)
            param_dict[f"{{{p.name}}}"] = display_val

        # 2. Lấy dữ liệu
        placeholders = ["%s::date" if p.param_type == 'date' else "%s::varchar" if p.param_type == 'char' else "%s::integer" for p in self.param_ids.sorted('sequence')]
        query = f"SELECT * FROM {self.store_procedure_name}({', '.join(placeholders)})"
        self.env.cr.execute(query, params)
        data_rows = self.env.cr.dictfetchall()

        # 3. Mở Excel
        try:
            wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.template_file)))
            sheet = wb.active
        except:
            raise UserError(_("File mẫu không hợp lệ."))
        
        # A. Replace Header Tags
        for r in range(1, self.start_row):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if isinstance(cell.value, str):
                    for tag, v in param_dict.items():
                        if tag in cell.value:
                            cell.value = cell.value.replace(tag, v)

        # B. Phân tích Mapping & Style
        col_mapping = {}
        cell_styles = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=self.start_row, column=col)
            if cell.value and isinstance(cell.value, str) and '{' in cell.value:
                key = cell.value.replace('{','').replace('}','').strip()
                col_mapping[key] = col
                cell_styles[col] = {
                    'font': copy(cell.font), 'border': copy(cell.border),
                    'fill': copy(cell.fill), 'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                }
        
        # Xóa dòng mẫu
        for col_idx in range(1, sheet.max_column + 1):
            sheet.cell(row=self.start_row, column=col_idx).value = ""

        # C. Đổ dữ liệu với Logic Group & Sum
        current_row = self.start_row
        group_cols = [gc.name for gc in self.group_by_column_ids]
        sum_cols = [sc.name for sc in self.sum_column_ids]
        last_group_values = [None] * len(group_cols)
        
        group_colors = ["D9EAD3", "C9DAF8", "EFEFEF", "F8F9FA", "FFF2CC"]
        bold_font = Font(bold=True)

        for idx, record in enumerate(data_rows):
            # Kiểm tra thay đổi cấp độ Group
            changed_level = -1
            for i, col in enumerate(group_cols):
                if record.get(col) != last_group_values[i]:
                    changed_level = i
                    break
            
            # Nếu có thay đổi, in Header Group mới kèm SUM
            if changed_level != -1:
                for i in range(changed_level, len(group_cols)):
                    col_name = group_cols[i]
                    val = record.get(col_name) or ''
                    
                    # LOGIC SUM: Có try...except để tránh lỗi ép kiểu dữ liệu chuỗi (như 'Mì dây 7')
                    current_group_sums = {sc: 0.0 for sc in sum_cols}
                    for scan_idx in range(idx, len(data_rows)):
                        scan_rec = data_rows[scan_idx]
                        is_same_group = True
                        for check_i in range(i + 1):
                            if scan_rec.get(group_cols[check_i]) != record.get(group_cols[check_i]):
                                is_same_group = False
                                break
                        
                        if is_same_group:
                            for sc in sum_cols:
                                try:
                                    current_group_sums[sc] += float(scan_rec.get(sc) or 0)
                                except (ValueError, TypeError):
                                    continue
                        else:
                            break

                    # In dòng Group Header (KHÔNG MERGE CELL)
                    indent = "    " * i
                    cell_label = sheet.cell(row=current_row, column=1)
                    cell_label.value = f"{indent}● {val}"

                    # Đổ giá trị SUM vào các cột tương ứng trên dòng Header
                    for sc, total_val in current_group_sums.items():
                        if sc in col_mapping:
                            c_idx = col_mapping[sc]
                            cell_sum = sheet.cell(row=current_row, column=c_idx)
                            cell_sum.value = total_val
                            # Áp dụng format số
                            if c_idx in cell_styles:
                                cell_sum.number_format = cell_styles[c_idx]['number_format']

                    # Style cho dòng Group (Tô màu nền và in đậm toàn dòng)
                    bg_color = group_colors[i] if i < len(group_colors) else "FFFFFF"
                    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    for c in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=current_row, column=c)
                        cell.fill = fill
                        cell.font = bold_font
                    
                    current_row += 1
                    last_group_values[i] = record.get(col_name)

            # Ghi dòng chi tiết
            for key, col_idx in col_mapping.items():
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = record.get(key, "")
                if col_idx in cell_styles:
                    s = cell_styles[col_idx]
                    cell.font, cell.border, cell.alignment, cell.number_format = copy(s['font']), copy(s['border']), copy(s['alignment']), s['number_format']
            
            current_row += 1

        # 4. Xuất file
        output = io.BytesIO()
        wb.save(output)
        out_data = base64.b64encode(output.getvalue())
        fname = f"{self.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': fname, 'datas': out_data, 'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        return {
            'type': 'ir.actions.act_url', 
            'url': f'/web/content/{attachment.id}?download=true', 
            'target': 'new'
        }

class ExcelReportColumn(models.Model):
    _name = 'excel.report.column'
    _description = 'Cột dữ liệu SQL'
    _rec_name = 'name'
    config_id = fields.Many2one('excel.report.config', ondelete='cascade')
    name = fields.Char(string='Tên cột')

class ExcelReportParam(models.Model):
    _name = 'excel.report.param'
    _description = 'Tham số báo cáo'
    _order = 'sequence'
    config_id = fields.Many2one('excel.report.config')
    sequence = fields.Integer(default=10)
    name = fields.Char(string='Mã tham số (tag)', required=True)
    display_name_param = fields.Char(string='Tên hiển thị', required=True)
    param_type = fields.Selection([
        ('date', 'Ngày'), 
        ('char', 'Chuỗi'), 
        ('integer', 'Số nguyên')
    ], default='date')
    value_date = fields.Date(string="Giá trị ngày")
    value_char = fields.Char(string="Giá trị chuỗi")
    value_integer = fields.Integer(string="Giá trị số")